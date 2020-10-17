from openpyxl import load_workbook
from openpyxl import Workbook

# 获取工作表
path = 'C:/Users/Administrator/Desktop/'

rows = ['订单号','商品名称','商品规格','成交数量','收货人姓名','收货人电话','收货地址']
data = {}
for row in rows:
  data[row] = []

files = [
  '快手小店批量导出-2020-10-17+09_38',
  '快手小店批量导出-2020-10-17+09_40 (1)',
  '快手小店批量导出-2020-10-17+09_40 (2)',
  '快手小店批量导出-2020-10-17+09_40 (3)',
  '快手小店批量导出-2020-10-17+09_40 (4)',
  '快手小店批量导出-2020-10-17+09_40 (5)',
  '快手小店批量导出-2020-10-17+09_40'
  ]

subfix = '.xlsx'

for file in files:
  rwb = load_workbook(path + file + subfix)
  rws = rwb.active

  for i in range(1, rws.max_column+1):
    for row in rows:
      if rws.cell(1, i).value == row:
        for j in range(2, rws.max_row+1):
          data[row].append(rws.cell(j, i).value)

# 写文件
wwb = Workbook()
wws = wwb.active

for index, row in enumerate(rows):
  wws.cell(1,index+1).value = row

for i in range(2, len(data[rows[1]])+2):
  for index, row in enumerate(rows):
    wws.cell(i,index+1).value = data[row][i-2]

wwb.save('C:/Users/Administrator/Desktop/wwb.xlsx')

