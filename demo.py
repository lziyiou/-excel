from openpyxl import load_workbook
from openpyxl import Workbook

# 获取工作表
path = 'C:/Users/Administrator/Desktop/'

data = {}
data['编号']=[]
data['直播价']=[]

files = []
files.append('a')
files.append('b')
files.append('c')

subfix = '.xlsx'

for file in files:
	rwb = load_workbook(path + file + subfix)
	rws = rwb.active

	for i in range(1, rws.max_column+1):
		if rws.cell(1, i).value == '编号':
			for j in range(2, rws.max_row+1):
				data['编号'].append(rws.cell(j, i).value)
		if rws.cell(1, i).value == '直播价':
			for j in range(2, rws.max_row+1):
				data['直播价'].append(rws.cell(j, i).value)

# 写文件
wwb = Workbook()
wws = wwb.active

wws.cell(1,1).value = '编号'
wws.cell(1,2).value = '直播价'

for i in range(2, len(data['编号'])+2):
    wws.cell(i, 1).value = data['编号'][i-2]
    wws.cell(i, 2).value = data['直播价'][i-2]

wwb.save('C:/Users/Administrator/Desktop/wwb.xlsx')

